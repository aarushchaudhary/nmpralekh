import logging
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell import WriteOnlyCell
from datetime import date, datetime
from celery import shared_task, group
from django.conf import settings

logger = logging.getLogger(__name__)

EXPORT_DIR = os.path.join(settings.BASE_DIR, 'exports')


def ensure_export_dir():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def _append_styled_headers(ws, headers):
    """Write a bold + blue header row in write-only mode."""
    styled_row = []
    for header in headers:
        cell           = WriteOnlyCell(ws, value=header)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center')
        styled_row.append(cell)
    ws.append(styled_row)


def build_campus_workbook(school_ids, date_from=None, date_to=None):
    """
    Builds a streaming write-only workbook for the given school IDs.
    Uses write_only=True so openpyxl never holds the full DOM in RAM,
    and .iterator() so Django streams DB rows one-by-one.

    Args:
        school_ids: list of school PKs to include.
        date_from:  optional lower bound (inclusive) applied to each sheet's
                    primary date field.  Passed as a date/str understood by
                    Django's ORM (``gte`` lookup).
        date_to:    optional upper bound (inclusive) applied to each sheet's
                    primary date field.  Passed as a date/str understood by
                    Django's ORM (``lte`` lookup).
    """
    from apps.records.models import (
        SchoolActivity, StudentActivity,
        FacultyFDPWorkshopGL, FacultyPublication,
        Patent, Certification, PlacementActivity
    )

    # write_only=True: rows are written directly to the ZIP stream,
    # never accumulated in memory as a full DOM tree.
    wb = openpyxl.Workbook(write_only=True)
    total_records = 0

    # Build reusable date-range filter kwargs for each date-field name.
    def _date_filters(field):
        kwargs = {}
        if date_from:
            kwargs[f'{field}__gte'] = date_from
        if date_to:
            kwargs[f'{field}__lte'] = date_to
        return kwargs

    sheets = [
        (
            'School Activities',
            SchoolActivity.objects.filter(
                school_id__in=school_ids, is_deleted=False,
                **_date_filters('date')
            ).select_related('school'),
            ['School', 'Name', 'Date', 'Details', 'School Wide'],
            lambda r: [
                r.school.name, r.name, str(r.date),
                r.details, 'Yes' if r.is_school_wide else 'No',
            ]
        ),
        (
            'Student Activities',
            StudentActivity.objects.filter(
                school_id__in=school_ids, is_deleted=False,
                **_date_filters('date')
            ).select_related('school'),
            ['School', 'Name', 'Date', 'Details', 'Conducted By', 'Type'],
            lambda r: [
                r.school.name, r.name, str(r.date), r.details,
                r.club_name or r.conducted_by or '',
                r.activity_type,
            ]
        ),
        (
            'FDP Workshop GL',
            FacultyFDPWorkshopGL.objects.filter(
                school_id__in=school_ids, is_deleted=False,
                **_date_filters('date_start')  # FDP uses date_start as primary date
            ).select_related('school'),
            ['School', 'Faculty', 'Name', 'Type',
             'Date Start', 'Date End', 'Organizing Body'],
            lambda r: [
                r.school.name, r.faculty_name, r.name, r.type,
                str(r.date_start),
                str(r.date_end) if r.date_end else '',
                r.organizing_body or '',
            ]
        ),
        (
            'Publications',
            FacultyPublication.objects.filter(
                school_id__in=school_ids, is_deleted=False,
                **_date_filters('date')
            ).select_related('school').prefetch_related('authors'),
            ['School', 'Author', 'Co-Authors', 'Title', 'Journal',
             'Date', 'Venue', 'Publication'],
            lambda r: [
                r.school.name, r.author_name,
                ", ".join(a.name for a in r.authors.all()),
                r.title_of_paper,
                r.journal_or_conference_name, str(r.date),
                r.venue or '', r.publication or '',
            ]
        ),
        (
            'Patents',
            Patent.objects.filter(
                school_id__in=school_ids, is_deleted=False,
                **_date_filters('date_of_publication')  # Patent uses date_of_publication
            ).select_related('school').prefetch_related('applicants'),
            ['School', 'Applicant', 'Co-Applicants', 'Title', 'Date',
             'Journal No', 'Status'],
            lambda r: [
                r.school.name, r.applicant_name,
                ", ".join(a.name for a in r.applicants.all()),
                r.title_of_patent,
                str(r.date_of_publication),
                r.journal_number, r.patent_status,
            ]
        ),
        (
            'Certifications',
            Certification.objects.filter(
                school_id__in=school_ids, is_deleted=False,
                **_date_filters('date')
            ).select_related('school'),
            ['School', 'Name', 'Course', 'Agency', 'Date', 'Link'],
            lambda r: [
                r.school.name, r.name, r.title_of_course,
                r.agency, str(r.date),
                r.credly_or_proof_link or '',
            ]
        ),
        (
            'Placements',
            PlacementActivity.objects.filter(
                school_id__in=school_ids, is_deleted=False,
                **_date_filters('date')
            ).select_related('school'),
            ['School', 'Name', 'Date', 'Details', 'Company'],
            lambda r: [
                r.school.name, r.name, str(r.date),
                r.details, r.company_name or '',
            ]
        ),
    ]

    for sheet_title, queryset, headers, row_fn in sheets:
        ws = wb.create_sheet(title=sheet_title)
        _append_styled_headers(ws, headers)
        count = 0
        # .iterator() streams rows from Postgres one-by-one instead of
        # loading the entire result set into a Python list.
        for record in queryset.iterator(chunk_size=2000):
            try:
                ws.append(row_fn(record))
                count += 1
            except Exception as e:
                logger.warning(
                    f"Skipped record id={record.id} "
                    f"in sheet '{sheet_title}' during export: {e}"
                )
                continue
        total_records += count

    return wb, total_records


@shared_task(
    bind=True,
    name='apps.export.tasks.export_single_campus',
    max_retries=3,
    default_retry_delay=60,
    autoretry_for=(Exception,),
)
def export_single_campus(self, campus_id):
    """
    Exports a single active campus to an Excel workbook and saves the
    resulting file + database record.  Retried automatically up to 3 times
    (with a 60-second delay) on any exception.
    """
    from apps.schools.models import Campus, School
    from apps.export.models import GeneratedExport

    ensure_export_dir()
    today = date.today().strftime('%Y-%m-%d')

    campus = Campus.objects.get(pk=campus_id)
    school_ids = list(
        School.objects.filter(
            campus=campus, is_active=True
        ).values_list('id', flat=True)
    )

    if not school_ids:
        return f'No active schools for campus {campus.name} — skipped'

    wb, total_records = build_campus_workbook(school_ids)

    filename = f'{campus.code}_MIS_Export_{today}.xlsx'
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    file_size_kb = os.path.getsize(filepath) // 1024

    export = GeneratedExport.objects.create(
        campus       = campus,
        export_type  = 'nightly',
        filename     = filename,
        filepath     = filepath,
        file_size_kb = file_size_kb,
        record_count = total_records,
    )

    return {
        'export_id':   export.id,
        'campus':      campus.name,
        'filename':    filename,
        'records':     total_records,
    }


@shared_task(name='apps.export.tasks.generate_nightly_exports')
def generate_nightly_exports():
    """
    Runs every night at 12:00 AM IST.
    Dispatches one export_single_campus task per active campus in parallel
    using celery.group so that a failure in one campus does not affect others.
    """
    from apps.schools.models import Campus

    campus_ids = list(
        Campus.objects.filter(is_active=True).values_list('id', flat=True)
    )

    if not campus_ids:
        return 'No active campuses found — nothing to export'

    job = group(export_single_campus.s(cid) for cid in campus_ids)
    job.apply_async()

    return f'Dispatched export jobs for {len(campus_ids)} campus(es)'


@shared_task(name='apps.export.tasks.generate_manual_export')
def generate_manual_export(campus_id, school_ids,
                            filters=None, user_id=None):
    """
    Triggered manually by master admin from the portal.
    """
    filters = filters or {}
    from apps.schools.models import Campus
    from apps.export.models import GeneratedExport
    from apps.accounts.models import User

    ensure_export_dir()

    campus   = Campus.objects.get(pk=campus_id)
    today    = date.today().strftime('%Y-%m-%d')
    now      = datetime.now().strftime('%H%M%S')

    wb, total_records = build_campus_workbook(
        school_ids,
        date_from=filters.get('date_from') or None,
        date_to=filters.get('date_to')   or None,
    )

    filename = f'{campus.code}_Manual_Export_{today}_{now}.xlsx'
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    file_size_kb = os.path.getsize(filepath) // 1024

    user = None
    if user_id:
        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            pass  # user deactivated between trigger and task execution
                  # export still proceeds, generated_by will be null

    export = GeneratedExport.objects.create(
        campus          = campus,
        export_type     = 'manual',
        filename        = filename,
        filepath        = filepath,
        generated_by    = user,
        file_size_kb    = file_size_kb,
        record_count    = total_records,
        date_range_from = filters.get('date_from') or None,
        date_range_to   = filters.get('date_to')   or None,
    )

    return {
        'export_id': export.id,
        'filename':  filename,
        'campus':    campus.name,
        'records':   total_records,
    }
