import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated

from apps.accounts.permissions import IsAdminOrUserOrSuperAdmin
from apps.schools.utils import get_user_school_ids
from apps.records.models import (
    SchoolActivity, StudentActivity,
    FacultyFDPWorkshopGL, FacultyPublication,
    Patent, Certification, PlacementActivity,
)


def style_header_row(ws, headers):
    """Applies bold + blue background to header row"""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center')


def get_scoped_queryset(model, user):
    school_ids = get_user_school_ids(user)
    return model.objects.filter(
        school_id__in=school_ids,
        is_deleted=False
    ).select_related('school')


def build_apply_filters(school_id, date_from, date_to):
    """Returns a reusable filter function"""
    def apply_filters(qs):
        from django.core.exceptions import FieldError
        try:
            if school_id: qs = qs.filter(school_id=school_id)
            if date_from: qs = qs.filter(date__gte=date_from)
            if date_to:   qs = qs.filter(date__lte=date_to)
        except FieldError:
            pass

        try:
            if date_from: qs = qs.filter(date_start__gte=date_from)
            if date_to:   qs = qs.filter(date_start__lte=date_to)
        except FieldError: pass

        try:
            if date_from: qs = qs.filter(date_of_publication__gte=date_from)
            if date_to:   qs = qs.filter(date_of_publication__lte=date_to)
        except FieldError: pass

        return qs[:5000]   # hard safety cap
    return apply_filters


class ExportSchoolActivitiesView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_id = request.query_params.get('school_id')
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        apply_filters = build_apply_filters(school_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'School Activity'

        headers = ['School', 'Name', 'Date', 'Details', 'School Wide']
        style_header_row(ws, headers)

        for act in apply_filters(get_scoped_queryset(SchoolActivity, request.user)):
            ws.append([
                act.school.name,
                act.name,
                str(act.date),
                act.details,
                'Yes' if act.is_school_wide else 'No',
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=school_activities.xlsx'
        wb.save(response)
        return response


class ExportStudentActivitiesView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_id = request.query_params.get('school_id')
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        apply_filters = build_apply_filters(school_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Student Activity'

        headers = ['School', 'Name', 'Date', 'Details', 'Conducted By', 'Type']
        style_header_row(ws, headers)

        for act in apply_filters(get_scoped_queryset(StudentActivity, request.user)):
            ws.append([
                act.school.name,
                act.name,
                str(act.date),
                act.details,
                act.conducted_by,
                act.activity_type,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=student_activities.xlsx'
        wb.save(response)
        return response


class ExportFDPView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_id = request.query_params.get('school_id')
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        apply_filters = build_apply_filters(school_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Faculty FDP, Workshop, GL'

        headers = ['School', 'Faculty Name', 'Date Start', 'Date End', 'Name', 'Details', 'Type', 'Organizing Body']
        style_header_row(ws, headers)

        for fdp in apply_filters(get_scoped_queryset(FacultyFDPWorkshopGL, request.user)):
            ws.append([
                fdp.school.name,
                fdp.faculty_name,
                str(fdp.date_start),
                str(fdp.date_end) if fdp.date_end else '',
                fdp.name,
                fdp.details,
                fdp.type,
                fdp.organizing_body or '',
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=faculty_fdp.xlsx'
        wb.save(response)
        return response


class ExportPublicationsView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_id = request.query_params.get('school_id')
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        apply_filters = build_apply_filters(school_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Faculty Publication'

        headers = ['School', 'Author Name', 'Author Type', 'Title of Paper',
                   'Journal/Conference', 'Date', 'Venue', 'Publication', 'DOI/Link']
        style_header_row(ws, headers)

        for pub in apply_filters(get_scoped_queryset(FacultyPublication, request.user)):
            ws.append([
                pub.school.name,
                pub.author_name,
                pub.author_type,
                pub.title_of_paper,
                pub.journal_or_conference_name,
                str(pub.date),
                pub.venue or '',
                pub.publication or '',
                pub.doi_or_link or '',
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=publications.xlsx'
        wb.save(response)
        return response


class ExportPatentsView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_id = request.query_params.get('school_id')
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        apply_filters = build_apply_filters(school_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PATENT'

        headers = ['School', 'Applicant Name', 'Applicant Type', 'Title of Patent',
                   'Details', 'Date of Publication', 'Journal Number', 'Status']
        style_header_row(ws, headers)

        for patent in apply_filters(get_scoped_queryset(Patent, request.user)):
            ws.append([
                patent.school.name,
                patent.applicant_name,
                patent.applicant_type,
                patent.title_of_patent,
                patent.details or '',
                str(patent.date_of_publication),
                patent.journal_number,
                patent.patent_status,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=patents.xlsx'
        wb.save(response)
        return response


class ExportCertificationsView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_id = request.query_params.get('school_id')
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        apply_filters = build_apply_filters(school_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'CERTIFICATES'

        headers = ['School', 'Date', 'Name', 'Title of Course',
                   'Details', 'Agency', 'Proof Link', 'Person Type']
        style_header_row(ws, headers)

        for cert in apply_filters(get_scoped_queryset(Certification, request.user)):
            ws.append([
                cert.school.name,
                str(cert.date),
                cert.name,
                cert.title_of_course,
                cert.details or '',
                cert.agency,
                cert.credly_or_proof_link or '',
                cert.person_type,
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=certifications.xlsx'
        wb.save(response)
        return response


class ExportPlacementsView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_id = request.query_params.get('school_id')
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        apply_filters = build_apply_filters(school_id, date_from, date_to)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Placement Activities'

        headers = ['School', 'Name', 'Date', 'Details', 'Company Name']
        style_header_row(ws, headers)

        for placement in apply_filters(get_scoped_queryset(PlacementActivity, request.user)):
            ws.append([
                placement.school.name,
                placement.name,
                str(placement.date),
                placement.details,
                placement.company_name or '',
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=placements.xlsx'
        wb.save(response)
        return response


from .tasks import build_campus_workbook

class ExportAllView(APIView):
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request):
        school_ids = list(get_user_school_ids(request.user))

        wb, total_records = build_campus_workbook(school_ids)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=MIS_Dashboard.xlsx'
        wb.save(response)
        return response


class ExportStatusView(APIView):
    """Frontend polls this every 2 seconds until ready"""
    permission_classes = [IsAdminOrUserOrSuperAdmin]

    def get(self, request, task_id):
        from celery.result import AsyncResult
        from django.core.cache import cache
        from django.http import HttpResponse
        from rest_framework.response import Response
        
        result    = AsyncResult(task_id)
        cache_key = f'export_{task_id}'
        file_data = cache.get(cache_key)

        if file_data:
            # file is ready — send it
            response = HttpResponse(
                file_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=MIS_Dashboard.xlsx'
            return response

        if result.failed():
            return Response({'status': 'failed'}, status=500)

        return Response({'status': 'processing'})


from rest_framework.response import Response
from apps.export.models import GeneratedExport
from apps.export.tasks import (
    generate_nightly_exports,
    generate_manual_export
)
from apps.accounts.permissions import IsMaster


class ExportHistoryView(APIView):
    permission_classes = [IsMaster]

    def get(self, request):
        exports   = GeneratedExport.objects.select_related(
                        'campus', 'generated_by'
                    )
        campus_id = request.query_params.get('campus_id')
        if campus_id:
            exports = exports.filter(campus_id=campus_id)

        data = [{
            'id':             e.id,
            'campus':         e.campus.name if e.campus else 'All',
            'campus_code':    e.campus.code if e.campus else 'ALL',
            'export_type':    e.export_type,
            'filename':       e.filename,
            'generated_by':   e.generated_by.full_name if e.generated_by else 'System',
            'generated_at':   e.generated_at.isoformat(),
            'file_size_kb':   e.file_size_kb,
            'record_count':   e.record_count,
            'date_range_from': str(e.date_range_from) if e.date_range_from else None,
            'date_range_to':   str(e.date_range_to)   if e.date_range_to   else None,
        } for e in exports]

        return Response(data)


class ExportDownloadView(APIView):
    permission_classes = [IsMaster]

    def get(self, request, pk):
        try:
            export = GeneratedExport.objects.get(pk=pk)
        except GeneratedExport.DoesNotExist:
            return Response({'detail': 'Export not found'}, status=404)

        if not os.path.exists(export.filepath):
            return Response(
                {'detail': 'File no longer exists on server'},
                status=404
            )

        with open(export.filepath, 'rb') as f:
            response = HttpResponse(
                f.read(),
                content_type=(
                    'application/vnd.openxmlformats-officedocument'
                    '.spreadsheetml.sheet'
                )
            )
            response['Content-Disposition'] = (
                f'attachment; filename="{export.filename}"'
            )
            return response


class TriggerManualExportView(APIView):
    permission_classes = [IsMaster]

    def post(self, request):
        campus_id = request.data.get('campus_id')
        date_from = request.data.get('date_from')
        date_to   = request.data.get('date_to')

        if not campus_id:
            return Response(
                {'detail': 'campus_id is required'},
                status=400
            )

        from apps.schools.models import School
        school_ids = list(
            School.objects.filter(
                campus_id=campus_id,
                is_active=True
            ).values_list('id', flat=True)
        )

        task = generate_manual_export.delay(
            campus_id  = campus_id,
            school_ids = school_ids,
            filters    = {
                'date_from': date_from,
                'date_to':   date_to,
            },
            user_id = request.user.id
        )

        return Response({
            'detail':  'Export started in background',
            'task_id': task.id,
        }, status=202)


class TriggerNightlyExportView(APIView):
    permission_classes = [IsMaster]

    def post(self, request):
        task = generate_nightly_exports.delay()
        return Response({
            'detail':  'Nightly export triggered manually',
            'task_id': task.id,
        }, status=202)


class ExportTaskStatusView(APIView):
    """Frontend polls this to check if export is ready"""
    permission_classes = [IsMaster]

    def get(self, request, task_id):
        from celery.result import AsyncResult
        result = AsyncResult(task_id)

        if result.successful():
            return Response({
                'status': 'completed',
                'result': result.result,
            })
        elif result.failed():
            return Response({'status': 'failed'}, status=500)
        else:
            return Response({'status': 'processing'})


# ─────────────────────────────────────────────
# MIS COORDINATOR EXPORT
# ─────────────────────────────────────────────
import json as json_module
from apps.accounts.permissions import IsMISCoordinator
from apps.records.serializers import (
    SchoolActivitySerializer,
    StudentActivitySerializer,
    FacultyFDPWorkshopGLSerializer,
    FacultyPublicationSerializer,
    PatentSerializer,
    CertificationSerializer,
    PlacementActivitySerializer,
)


def _gather_coordinator_data(school_ids, date_from=None, date_to=None):
    """
    Queries all 7 MIS record modules for the given school IDs
    with optional date range filtering.
    Returns a list of (module_name, queryset, headers, row_fn) tuples for Excel,
    and also the raw querysets for JSON serialization.
    """
    from apps.records.models import (
        SchoolActivity, StudentActivity,
        FacultyFDPWorkshopGL, FacultyPublication,
        Patent, Certification, PlacementActivity,
    )

    def _filter_qs(qs, date_field, d_from, d_to):
        if d_from:
            qs = qs.filter(**{f'{date_field}__gte': d_from})
        if d_to:
            qs = qs.filter(**{f'{date_field}__lte': d_to})
        return qs

    base = dict(school_id__in=school_ids, is_deleted=False)

    modules = [
        {
            'name': 'School Activities',
            'model': SchoolActivity,
            'date_field': 'date',
            'qs': SchoolActivity.objects.filter(**base).select_related('school'),
            'headers': ['School', 'Name', 'Date', 'Details', 'School Wide'],
            'row_fn': lambda r: [
                r.school.name, r.name, str(r.date),
                r.details, 'Yes' if r.is_school_wide else 'No',
            ],
            'serializer': SchoolActivitySerializer,
        },
        {
            'name': 'Student Activities',
            'model': StudentActivity,
            'date_field': 'date',
            'qs': StudentActivity.objects.filter(**base).select_related('school'),
            'headers': ['School', 'Name', 'Date', 'Details', 'Conducted By', 'Type'],
            'row_fn': lambda r: [
                r.school.name, r.name, str(r.date), r.details,
                r.club_name or r.conducted_by or '', r.activity_type,
            ],
            'serializer': StudentActivitySerializer,
        },
        {
            'name': 'FDP Workshop GL',
            'model': FacultyFDPWorkshopGL,
            'date_field': 'date_start',
            'qs': FacultyFDPWorkshopGL.objects.filter(**base).select_related('school'),
            'headers': ['School', 'Faculty', 'Name', 'Type',
                        'Date Start', 'Date End', 'Organizing Body'],
            'row_fn': lambda r: [
                r.school.name, r.faculty_name, r.name, r.type,
                str(r.date_start),
                str(r.date_end) if r.date_end else '',
                r.organizing_body or '',
            ],
            'serializer': FacultyFDPWorkshopGLSerializer,
        },
        {
            'name': 'Publications',
            'model': FacultyPublication,
            'date_field': 'date',
            'qs': FacultyPublication.objects.filter(**base).select_related('school', 'created_by').prefetch_related('authors'),
            'headers': ['School', 'Author', 'Type', 'Title', 'Journal',
                        'Date', 'Venue', 'Publication', 'DOI/Link'],
            'row_fn': lambda r: [
                r.school.name, r.author_name, r.author_type,
                r.title_of_paper, r.journal_or_conference_name,
                str(r.date), r.venue or '', r.publication or '',
                r.doi_or_link or '',
            ],
            'serializer': FacultyPublicationSerializer,
        },
        {
            'name': 'Patents',
            'model': Patent,
            'date_field': 'date_of_publication',
            'qs': Patent.objects.filter(**base).select_related('school', 'created_by').prefetch_related('applicants'),
            'headers': ['School', 'Applicant', 'Type', 'Title',
                        'Date', 'Journal No', 'Status', 'DOI/Link'],
            'row_fn': lambda r: [
                r.school.name, r.applicant_name, r.applicant_type,
                r.title_of_patent, str(r.date_of_publication),
                r.journal_number, r.patent_status,
                r.doi_or_link or '',
            ],
            'serializer': PatentSerializer,
        },
        {
            'name': 'Certifications',
            'model': Certification,
            'date_field': 'date',
            'qs': Certification.objects.filter(**base).select_related('school', 'created_by'),
            'headers': ['School', 'Name', 'Course', 'Agency',
                        'Date', 'Proof Link', 'Person Type'],
            'row_fn': lambda r: [
                r.school.name, r.name, r.title_of_course,
                r.agency, str(r.date),
                r.credly_or_proof_link or '', r.person_type,
            ],
            'serializer': CertificationSerializer,
        },
        {
            'name': 'Placements',
            'model': PlacementActivity,
            'date_field': 'date',
            'qs': PlacementActivity.objects.filter(**base).select_related('school'),
            'headers': ['School', 'Name', 'Date', 'Details', 'Company'],
            'row_fn': lambda r: [
                r.school.name, r.name, str(r.date),
                r.details, r.company_name or '',
            ],
            'serializer': PlacementActivitySerializer,
        },
    ]

    # apply date filtering
    for mod in modules:
        mod['qs'] = _filter_qs(mod['qs'], mod['date_field'], date_from, date_to)

    return modules


class CoordinatorExportView(APIView):
    """
    GET /api/export/coordinator/
    Query params: date_from, date_to, format (excel|json)
    Protected: IsMISCoordinator only
    """
    permission_classes = [IsMISCoordinator]

    def perform_content_negotiation(self, request, force=False):
        # Ignore DRF's format override so 'format=excel' doesn't cause 404
        from rest_framework.renderers import JSONRenderer
        return (JSONRenderer(), JSONRenderer.media_type)

    def get(self, request):
        date_from = request.query_params.get('date_from')
        date_to   = request.query_params.get('date_to')
        fmt       = request.query_params.get('format', 'excel').lower()

        if fmt not in ('excel', 'json'):
            return Response(
                {'detail': 'format must be "excel" or "json"'},
                status=400
            )

        school_ids = list(get_user_school_ids(request.user))
        if not school_ids:
            return Response(
                {'detail': 'No schools assigned to your account'},
                status=400
            )

        modules = _gather_coordinator_data(school_ids, date_from, date_to)

        if fmt == 'json':
            return self._export_json(modules)
        return self._export_excel(modules)

    def _export_excel(self, modules):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for mod in modules:
            ws = wb.create_sheet(title=mod['name'][:31])  # Excel limit
            style_header_row(ws, mod['headers'])
            for record in mod['qs'][:5000]:
                try:
                    ws.append(mod['row_fn'](record))
                except Exception:
                    continue

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=MIS_Coordinator_Export.xlsx'
        wb.save(response)
        return response

    def _export_json(self, modules):
        data = {}
        for mod in modules:
            serializer = mod['serializer'](mod['qs'][:5000], many=True)
            data[mod['name']] = serializer.data

        json_bytes = json_module.dumps(data, indent=2, default=str).encode('utf-8')

        response = HttpResponse(
            json_bytes,
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename=MIS_Coordinator_Export.json'
        return response
